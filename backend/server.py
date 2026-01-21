from fastapi import FastAPI, APIRouter, HTTPException, Query, BackgroundTasks
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import httpx
from io import BytesIO
import asyncio
from contextlib import asynccontextmanager

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Weather Underground config
WU_API_KEY = os.environ.get('WEATHER_UNDERGROUND_API_KEY', '')
WU_STATION_ID = os.environ.get('WEATHER_UNDERGROUND_STATION_ID', '')
WU_BASE_URL = "https://api.weather.com/v2/pws"

# AEMET config
AEMET_API_KEY = os.environ.get('AEMET_API_KEY', '')
AEMET_MUNICIPIO = os.environ.get('AEMET_MUNICIPIO', '23091')  # Villacarrillo
AEMET_BASE_URL = "https://opendata.aemet.es/opendata/api"

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Background task for auto-fetching
auto_fetch_task = None

async def fetch_weather_periodically():
    """Background task to fetch weather data every 5 minutes"""
    while True:
        try:
            await fetch_and_store_current_weather()
            logger.info("Auto-fetched weather data")
        except Exception as e:
            logger.error(f"Error in auto-fetch: {e}")
        await asyncio.sleep(300)  # 5 minutes

@asynccontextmanager
async def lifespan(app: FastAPI):
    global auto_fetch_task
    # Startup
    logger.info("Starting weather auto-fetch task")
    auto_fetch_task = asyncio.create_task(fetch_weather_periodically())
    yield
    # Shutdown
    if auto_fetch_task:
        auto_fetch_task.cancel()
    client.close()

# Create the main app
app = FastAPI(lifespan=lifespan)

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Models
class WeatherObservation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    station_id: str
    timestamp: datetime
    temp_c: Optional[float] = None
    temp_f: Optional[float] = None
    humidity: Optional[float] = None
    dewpoint_c: Optional[float] = None
    dewpoint_f: Optional[float] = None
    heat_index_c: Optional[float] = None
    heat_index_f: Optional[float] = None
    wind_chill_c: Optional[float] = None
    wind_chill_f: Optional[float] = None
    wind_speed_kph: Optional[float] = None
    wind_speed_mph: Optional[float] = None
    wind_gust_kph: Optional[float] = None
    wind_gust_mph: Optional[float] = None
    wind_dir: Optional[int] = None
    pressure_mb: Optional[float] = None
    pressure_in: Optional[float] = None
    precip_rate_mm: Optional[float] = None
    precip_rate_in: Optional[float] = None
    precip_total_mm: Optional[float] = None
    precip_total_in: Optional[float] = None
    uv: Optional[float] = None
    solar_radiation: Optional[float] = None

class DailySummary(BaseModel):
    date: str
    temp_max_c: Optional[float] = None
    temp_min_c: Optional[float] = None
    temp_avg_c: Optional[float] = None
    humidity_avg: Optional[float] = None
    wind_avg_kph: Optional[float] = None
    wind_gust_max_kph: Optional[float] = None
    precip_total_mm: Optional[float] = None
    observation_count: int = 0

class WeatherResponse(BaseModel):
    status: str
    data: Optional[Dict[str, Any]] = None
    message: Optional[str] = None

# Weather Underground API functions
async def fetch_current_from_wu() -> Optional[Dict[str, Any]]:
    """Fetch current conditions from Weather Underground API"""
    url = f"{WU_BASE_URL}/observations/current"
    params = {
        "stationId": WU_STATION_ID,
        "format": "json",
        "units": "m",
        "apiKey": WU_API_KEY
    }
    
    async with httpx.AsyncClient() as http_client:
        try:
            response = await http_client.get(url, params=params, timeout=15.0)
            response.raise_for_status()
            return response.json()
        except httpx.HTTPError as e:
            logger.error(f"Error fetching current conditions: {e}")
            return None

async def fetch_history_from_wu(date_str: str) -> Optional[Dict[str, Any]]:
    """Fetch historical data from Weather Underground API for a specific date"""
    url = f"{WU_BASE_URL}/observations/all/1day"
    params = {
        "stationId": WU_STATION_ID,
        "format": "json",
        "units": "m",
        "apiKey": WU_API_KEY,
        "date": date_str  # Format: YYYYMMDD
    }
    
    async with httpx.AsyncClient() as http_client:
        try:
            response = await http_client.get(url, params=params, timeout=30.0)
            response.raise_for_status()
            return response.json()
        except httpx.HTTPError as e:
            logger.error(f"Error fetching history for {date_str}: {e}")
            return None

def parse_wu_observation(obs: Dict[str, Any]) -> WeatherObservation:
    """Parse Weather Underground observation into our model"""
    metric = obs.get("metric", {})
    imperial = obs.get("imperial", {})
    
    # Parse timestamp
    obs_time = obs.get("obsTimeUtc", obs.get("obsTimeLocal", ""))
    try:
        if obs_time:
            timestamp = datetime.fromisoformat(obs_time.replace("Z", "+00:00"))
        else:
            timestamp = datetime.now(timezone.utc)
    except:
        timestamp = datetime.now(timezone.utc)
    
    return WeatherObservation(
        station_id=WU_STATION_ID,
        timestamp=timestamp,
        temp_c=metric.get("temp"),
        temp_f=imperial.get("temp"),
        humidity=obs.get("humidity"),
        dewpoint_c=metric.get("dewpt"),
        dewpoint_f=imperial.get("dewpt"),
        heat_index_c=metric.get("heatIndex"),
        heat_index_f=imperial.get("heatIndex"),
        wind_chill_c=metric.get("windChill"),
        wind_chill_f=imperial.get("windChill"),
        wind_speed_kph=metric.get("windSpeed"),
        wind_speed_mph=imperial.get("windSpeed"),
        wind_gust_kph=metric.get("windGust"),
        wind_gust_mph=imperial.get("windGust"),
        wind_dir=obs.get("winddir"),
        pressure_mb=metric.get("pressure"),
        pressure_in=imperial.get("pressure"),
        precip_rate_mm=metric.get("precipRate"),
        precip_rate_in=imperial.get("precipRate"),
        precip_total_mm=metric.get("precipTotal"),
        precip_total_in=imperial.get("precipTotal"),
        uv=obs.get("uv"),
        solar_radiation=obs.get("solarRadiation")
    )

async def fetch_and_store_current_weather() -> Optional[WeatherObservation]:
    """Fetch current weather and store in database"""
    data = await fetch_current_from_wu()
    if not data or "observations" not in data or len(data["observations"]) == 0:
        return None
    
    obs = data["observations"][0]
    weather = parse_wu_observation(obs)
    
    # Store in MongoDB
    doc = weather.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.observations.insert_one(doc)
    
    return weather

# API Routes
@api_router.get("/")
async def root():
    return {"message": "Weather Station API", "station_id": WU_STATION_ID}

@api_router.get("/weather/current")
async def get_current_weather():
    """Get current weather conditions"""
    weather = await fetch_and_store_current_weather()
    
    if not weather:
        # Try to get latest from database
        latest = await db.observations.find_one(
            {}, 
            {"_id": 0},
            sort=[("timestamp", -1)]
        )
        if latest:
            return {"status": "success", "data": latest, "source": "cache"}
        raise HTTPException(status_code=503, detail="Unable to fetch weather data")
    
    return {"status": "success", "data": weather.model_dump(), "source": "live"}

@api_router.get("/weather/history")
async def get_weather_history(
    start_date: str = Query(..., description="Start date YYYYMMDD"),
    end_date: str = Query(..., description="End date YYYYMMDD")
):
    """Get historical weather data for a date range"""
    try:
        start = datetime.strptime(start_date, "%Y%m%d")
        end = datetime.strptime(end_date, "%Y%m%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYYMMDD")
    
    if end < start:
        raise HTTPException(status_code=400, detail="End date must be after start date")
    
    if (end - start).days > 31:
        raise HTTPException(status_code=400, detail="Date range cannot exceed 31 days")
    
    all_observations = []
    current = start
    
    while current <= end:
        date_str = current.strftime("%Y%m%d")
        
        # First check database
        cached = await db.observations.find(
            {
                "timestamp": {
                    "$gte": current.isoformat(),
                    "$lt": (current + timedelta(days=1)).isoformat()
                }
            },
            {"_id": 0}
        ).to_list(1000)
        
        if cached:
            all_observations.extend(cached)
        else:
            # Fetch from API
            data = await fetch_history_from_wu(date_str)
            if data and "observations" in data:
                for obs in data["observations"]:
                    weather = parse_wu_observation(obs)
                    doc = weather.model_dump()
                    doc['timestamp'] = doc['timestamp'].isoformat()
                    all_observations.append(doc)
                    # Store in database
                    await db.observations.insert_one(doc)
        
        current += timedelta(days=1)
    
    # Sort by timestamp
    all_observations.sort(key=lambda x: x.get("timestamp", ""))
    
    return {
        "status": "success",
        "count": len(all_observations),
        "start_date": start_date,
        "end_date": end_date,
        "data": all_observations
    }

@api_router.get("/weather/last24h")
async def get_last_24_hours():
    """Get weather data for the last 24 hours"""
    now = datetime.now(timezone.utc)
    yesterday = now - timedelta(hours=24)
    
    # Get from database
    observations = await db.observations.find(
        {
            "timestamp": {
                "$gte": yesterday.isoformat(),
                "$lte": now.isoformat()
            }
        },
        {"_id": 0}
    ).sort("timestamp", 1).to_list(1000)
    
    # If no data, try to fetch from API
    if len(observations) < 10:
        today = now.strftime("%Y%m%d")
        data = await fetch_history_from_wu(today)
        if data and "observations" in data:
            for obs in data["observations"]:
                weather = parse_wu_observation(obs)
                doc = weather.model_dump()
                doc['timestamp'] = doc['timestamp'].isoformat()
                # Check if already exists
                exists = await db.observations.find_one({"id": doc["id"]})
                if not exists:
                    await db.observations.insert_one(doc)
                    observations.append(doc)
    
    observations.sort(key=lambda x: x.get("timestamp", ""))
    
    return {
        "status": "success",
        "count": len(observations),
        "data": observations
    }

@api_router.get("/weather/statistics")
async def get_weather_statistics(
    start_date: str = Query(..., description="Start date YYYYMMDD"),
    end_date: str = Query(..., description="End date YYYYMMDD")
):
    """Calculate weather statistics for a date range"""
    try:
        start = datetime.strptime(start_date, "%Y%m%d").replace(tzinfo=timezone.utc)
        end = datetime.strptime(end_date, "%Y%m%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYYMMDD")
    
    pipeline = [
        {
            "$match": {
                "timestamp": {
                    "$gte": start.isoformat(),
                    "$lte": end.isoformat()
                }
            }
        },
        {
            "$group": {
                "_id": None,
                "temp_max_c": {"$max": "$temp_c"},
                "temp_min_c": {"$min": "$temp_c"},
                "temp_avg_c": {"$avg": "$temp_c"},
                "humidity_avg": {"$avg": "$humidity"},
                "humidity_max": {"$max": "$humidity"},
                "humidity_min": {"$min": "$humidity"},
                "wind_avg_kph": {"$avg": "$wind_speed_kph"},
                "wind_max_kph": {"$max": "$wind_speed_kph"},
                "wind_gust_max_kph": {"$max": "$wind_gust_kph"},
                "pressure_avg_mb": {"$avg": "$pressure_mb"},
                "pressure_max_mb": {"$max": "$pressure_mb"},
                "pressure_min_mb": {"$min": "$pressure_mb"},
                "precip_total_mm": {"$max": "$precip_total_mm"},
                "uv_max": {"$max": "$uv"},
                "solar_max": {"$max": "$solar_radiation"},
                "observation_count": {"$sum": 1}
            }
        }
    ]
    
    result = await db.observations.aggregate(pipeline).to_list(1)
    
    if not result:
        return {
            "status": "success",
            "statistics": None,
            "message": "No data found for the specified period"
        }
    
    stats = result[0]
    del stats["_id"]
    
    # Round values
    for key in stats:
        if isinstance(stats[key], float):
            stats[key] = round(stats[key], 1)
    
    return {
        "status": "success",
        "start_date": start_date,
        "end_date": end_date,
        "statistics": stats
    }

@api_router.get("/weather/export/excel")
async def export_to_excel(
    start_date: str = Query(..., description="Start date YYYYMMDD"),
    end_date: str = Query(..., description="End date YYYYMMDD")
):
    """Export weather data to Excel file"""
    try:
        start = datetime.strptime(start_date, "%Y%m%d").replace(tzinfo=timezone.utc)
        end = datetime.strptime(end_date, "%Y%m%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYYMMDD")
    
    # Get data
    observations = await db.observations.find(
        {
            "timestamp": {
                "$gte": start.isoformat(),
                "$lte": end.isoformat()
            }
        },
        {"_id": 0}
    ).sort("timestamp", 1).to_list(10000)
    
    if not observations:
        raise HTTPException(status_code=404, detail="No data found for the specified period")
    
    # Create Excel file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Sheet 1: Raw Data
    ws_data = wb.active
    ws_data.title = "Datos Meteorológicos"
    
    # Headers
    headers = [
        "Fecha/Hora", "Temp (°C)", "Temp (°F)", "Humedad (%)", 
        "Punto Rocío (°C)", "Viento (km/h)", "Ráfaga (km/h)", 
        "Dirección Viento", "Presión (mb)", "Lluvia (mm)", 
        "UV", "Radiación Solar"
    ]
    
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    for row, obs in enumerate(observations, 2):
        ws_data.cell(row=row, column=1, value=obs.get("timestamp", ""))
        ws_data.cell(row=row, column=2, value=obs.get("temp_c"))
        ws_data.cell(row=row, column=3, value=obs.get("temp_f"))
        ws_data.cell(row=row, column=4, value=obs.get("humidity"))
        ws_data.cell(row=row, column=5, value=obs.get("dewpoint_c"))
        ws_data.cell(row=row, column=6, value=obs.get("wind_speed_kph"))
        ws_data.cell(row=row, column=7, value=obs.get("wind_gust_kph"))
        ws_data.cell(row=row, column=8, value=obs.get("wind_dir"))
        ws_data.cell(row=row, column=9, value=obs.get("pressure_mb"))
        ws_data.cell(row=row, column=10, value=obs.get("precip_total_mm"))
        ws_data.cell(row=row, column=11, value=obs.get("uv"))
        ws_data.cell(row=row, column=12, value=obs.get("solar_radiation"))
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws_data.column_dimensions[get_column_letter(col)].width = 15
    
    # Sheet 2: Summary Statistics
    ws_stats = wb.create_sheet(title="Resumen")
    
    # Calculate statistics
    temps = [o.get("temp_c") for o in observations if o.get("temp_c") is not None]
    humidity = [o.get("humidity") for o in observations if o.get("humidity") is not None]
    winds = [o.get("wind_speed_kph") for o in observations if o.get("wind_speed_kph") is not None]
    gusts = [o.get("wind_gust_kph") for o in observations if o.get("wind_gust_kph") is not None]
    precip = [o.get("precip_total_mm") for o in observations if o.get("precip_total_mm") is not None]
    
    stats_data = [
        ["RESUMEN METEOROLÓGICO", ""],
        ["Período", f"{start_date} - {end_date}"],
        ["Estación", WU_STATION_ID],
        ["Total Observaciones", len(observations)],
        ["", ""],
        ["TEMPERATURA", ""],
        ["Máxima (°C)", max(temps) if temps else "N/A"],
        ["Mínima (°C)", min(temps) if temps else "N/A"],
        ["Media (°C)", round(sum(temps)/len(temps), 1) if temps else "N/A"],
        ["", ""],
        ["HUMEDAD", ""],
        ["Media (%)", round(sum(humidity)/len(humidity), 1) if humidity else "N/A"],
        ["", ""],
        ["VIENTO", ""],
        ["Velocidad Media (km/h)", round(sum(winds)/len(winds), 1) if winds else "N/A"],
        ["Ráfaga Máxima (km/h)", max(gusts) if gusts else "N/A"],
        ["", ""],
        ["PRECIPITACIÓN", ""],
        ["Total (mm)", max(precip) if precip else 0],
    ]
    
    for row, (label, value) in enumerate(stats_data, 1):
        cell_label = ws_stats.cell(row=row, column=1, value=label)
        cell_value = ws_stats.cell(row=row, column=2, value=value)
        if label and not value:
            cell_label.font = Font(bold=True, size=12)
        cell_label.alignment = Alignment(horizontal='left')
        cell_value.alignment = Alignment(horizontal='right')
    
    ws_stats.column_dimensions['A'].width = 25
    ws_stats.column_dimensions['B'].width = 20
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"weather_data_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )

@api_router.get("/station/info")
async def get_station_info():
    """Get station information"""
    return {
        "station_id": WU_STATION_ID,
        "api_configured": bool(WU_API_KEY and WU_STATION_ID),
        "database": os.environ.get('DB_NAME', 'unknown')
    }

# AEMET API Functions
async def fetch_aemet_data(endpoint: str) -> Optional[Dict[str, Any]]:
    """Fetch data from AEMET API (two-step process)"""
    headers = {
        "api_key": AEMET_API_KEY,
        "Accept": "application/json"
    }
    
    async with httpx.AsyncClient() as http_client:
        try:
            # Step 1: Get the data URL
            response = await http_client.get(
                f"{AEMET_BASE_URL}{endpoint}",
                headers=headers,
                timeout=15.0
            )
            response.raise_for_status()
            data = response.json()
            
            if data.get("estado") != 200 or "datos" not in data:
                logger.error(f"AEMET API error: {data}")
                return None
            
            # Step 2: Fetch actual data from the URL
            data_url = data["datos"]
            data_response = await http_client.get(data_url, timeout=15.0)
            data_response.raise_for_status()
            return data_response.json()
            
        except httpx.HTTPError as e:
            logger.error(f"Error fetching AEMET data: {e}")
            return None
        except Exception as e:
            logger.error(f"Error parsing AEMET data: {e}")
            return None

@api_router.get("/aemet/forecast")
async def get_aemet_forecast():
    """Get AEMET weather forecast for Villacarrillo"""
    data = await fetch_aemet_data(f"/prediccion/especifica/municipio/diaria/{AEMET_MUNICIPIO}")
    
    if not data or len(data) == 0:
        return {
            "status": "error",
            "message": "No se pudo obtener el pronóstico de AEMET",
            "forecast": None
        }
    
    try:
        prediction = data[0]
        municipio = prediction.get("nombre", "Villacarrillo")
        provincia = prediction.get("provincia", "Jaén")
        elaborado = prediction.get("elaborado", "")
        
        # Parse forecast days
        forecast_days = []
        prediccion = prediction.get("prediccion", {})
        dias = prediccion.get("dia", [])
        
        for dia in dias[:5]:  # Next 5 days
            fecha = dia.get("fecha", "")
            
            # Get temperature
            temp_data = dia.get("temperatura", {})
            temp_max = temp_data.get("maxima") if isinstance(temp_data, dict) else None
            temp_min = temp_data.get("minima") if isinstance(temp_data, dict) else None
            
            # Get sky condition
            estado_cielo = dia.get("estadoCielo", [])
            cielo_desc = ""
            if estado_cielo and len(estado_cielo) > 0:
                # Get midday condition
                for ec in estado_cielo:
                    if ec.get("periodo") in ["12-24", "00-24"] or not ec.get("periodo"):
                        cielo_desc = ec.get("descripcion", "")
                        break
                if not cielo_desc and estado_cielo:
                    cielo_desc = estado_cielo[0].get("descripcion", "")
            
            # Get precipitation probability
            prob_precip = dia.get("probPrecipitacion", [])
            max_prob = 0
            for pp in prob_precip:
                if isinstance(pp, dict):
                    val = pp.get("value") or pp.get("valor", 0)
                    if val and int(val) > max_prob:
                        max_prob = int(val)
            
            # Get wind
            viento = dia.get("viento", [])
            wind_speed = None
            wind_dir = None
            for v in viento:
                if isinstance(v, dict) and v.get("velocidad"):
                    wind_speed = v.get("velocidad")
                    wind_dir = v.get("direccion")
                    break
            
            # Get humidity
            humedad = dia.get("humedadRelativa", {})
            hum_max = humedad.get("maxima") if isinstance(humedad, dict) else None
            hum_min = humedad.get("minima") if isinstance(humedad, dict) else None
            
            forecast_days.append({
                "fecha": fecha,
                "temp_max": temp_max,
                "temp_min": temp_min,
                "cielo": cielo_desc,
                "prob_precipitacion": max_prob,
                "viento_velocidad": wind_speed,
                "viento_direccion": wind_dir,
                "humedad_max": hum_max,
                "humedad_min": hum_min
            })
        
        return {
            "status": "success",
            "municipio": municipio,
            "provincia": provincia,
            "elaborado": elaborado,
            "forecast": forecast_days
        }
        
    except Exception as e:
        logger.error(f"Error parsing AEMET forecast: {e}")
        return {
            "status": "error",
            "message": str(e),
            "forecast": None
        }

@api_router.get("/aemet/alerts")
async def get_aemet_alerts():
    """Get AEMET weather alerts for Andalucía/Jaén"""
    # Use CAP alerts for the region
    headers = {
        "api_key": AEMET_API_KEY,
        "Accept": "application/json"
    }
    
    async with httpx.AsyncClient() as http_client:
        try:
            # Get alerts for Andalucía (zone 61)
            response = await http_client.get(
                f"{AEMET_BASE_URL}/avisos_cap/ultimoelaborado/area/61",
                headers=headers,
                timeout=15.0
            )
            
            if response.status_code == 404:
                # No alerts
                return {
                    "status": "success",
                    "alerts": [],
                    "message": "No hay alertas activas"
                }
            
            response.raise_for_status()
            data = response.json()
            
            if data.get("estado") == 404 or "datos" not in data:
                return {
                    "status": "success",
                    "alerts": [],
                    "message": "No hay alertas activas"
                }
            
            # Fetch actual alert data
            alerts_url = data.get("datos")
            if not alerts_url:
                return {
                    "status": "success",
                    "alerts": [],
                    "message": "No hay alertas activas"
                }
            
            alerts_response = await http_client.get(alerts_url, timeout=15.0)
            
            # Parse XML alerts or return raw
            alerts_text = alerts_response.text
            
            # Try to parse as JSON first
            try:
                alerts_data = alerts_response.json()
                return {
                    "status": "success",
                    "alerts": alerts_data if isinstance(alerts_data, list) else [alerts_data],
                    "message": None
                }
            except:
                # It's likely XML CAP format - extract basic info
                import re
                
                alerts = []
                # Extract event types and descriptions from XML
                events = re.findall(r'<event>(.*?)</event>', alerts_text, re.DOTALL)
                headlines = re.findall(r'<headline>(.*?)</headline>', alerts_text, re.DOTALL)
                descriptions = re.findall(r'<description>(.*?)</description>', alerts_text, re.DOTALL)
                severities = re.findall(r'<severity>(.*?)</severity>', alerts_text, re.DOTALL)
                
                for i in range(len(events)):
                    alert = {
                        "event": events[i] if i < len(events) else "",
                        "headline": headlines[i] if i < len(headlines) else "",
                        "description": descriptions[i] if i < len(descriptions) else "",
                        "severity": severities[i] if i < len(severities) else "Unknown"
                    }
                    # Filter for Jaén if possible
                    if "Jaén" in alert.get("description", "") or "Jaén" in alert.get("headline", "") or not alerts:
                        alerts.append(alert)
                
                return {
                    "status": "success",
                    "alerts": alerts[:5],  # Max 5 alerts
                    "message": None if alerts else "No hay alertas activas para la zona"
                }
                
        except httpx.HTTPError as e:
            logger.error(f"Error fetching AEMET alerts: {e}")
            return {
                "status": "error",
                "alerts": [],
                "message": "Error al obtener alertas de AEMET"
            }

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
